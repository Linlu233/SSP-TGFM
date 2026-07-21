#!/usr/bin/env python3
"""Synchronize the complete experiment-result appendix into Markdown.

The workbook is the structured source of truth.  The main body of the report
stays concise, while this script adds every scientifically relevant result row
that was previously available only in the workbook.
"""

from __future__ import annotations

import json
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REPORT = ROOT / "EXPERIMENT_RESULTS.md"
WORKBOOK = ROOT / "EXPERIMENT_RESULTS.xlsx"

BEGIN = "<!-- BEGIN GENERATED COMPLETE APPENDIX -->"
END = "<!-- END GENERATED COMPLETE APPENDIX -->"

Column = tuple[str, str | Callable[[dict[str, Any]], Any]]


def sheet_records(workbook: Any, name: str) -> list[dict[str, Any]]:
    sheet = workbook[name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows]


def format_value(value: Any) -> str:
    if value is None or value == "":
        return "-"
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, float):
        if value != value:
            return "-"
        if abs(value) >= 100_000:
            return f"{value:.0f}"
        return f"{value:.6f}".rstrip("0").rstrip(".")
    if isinstance(value, (list, dict, tuple)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    text = str(value).replace("\r", " ").replace("\n", "<br>")
    return text.replace("|", "\\|")


def render_table(rows: Iterable[dict[str, Any]], columns: list[Column]) -> list[str]:
    output = [
        "| " + " | ".join(name for name, _ in columns) + " |",
        "| " + " | ".join("---" for _ in columns) + " |",
    ]
    for row in rows:
        values = []
        for _, source in columns:
            value = source(row) if callable(source) else row.get(source)
            values.append(format_value(value))
        output.append("| " + " | ".join(values) + " |")
    return output


def details(title: str, body: list[str]) -> list[str]:
    return ["<details>", f"<summary>{title}</summary>", "", *body, "", "</details>"]


def result_key(row: dict[str, Any], fields: list[str]) -> tuple[Any, ...]:
    return tuple(row.get(field) for field in fields)


def nonempty(*keys: str) -> Callable[[dict[str, Any]], Any]:
    def select(row: dict[str, Any]) -> Any:
        for key in keys:
            value = row.get(key)
            if value is not None and value != "":
                return value
        return None

    return select


def build_appendix(workbook: Any) -> tuple[str, dict[str, int]]:
    summaries = sheet_records(workbook, "All_Summaries")
    raw = sheet_records(workbook, "All_Results_Raw")
    partial = sheet_records(workbook, "Partial_Raw")
    search_runs = sheet_records(workbook, "Search_Runs")
    search_summaries = sheet_records(workbook, "Search_Summary")
    search_json = sheet_records(workbook, "Search_JSON_Rows")
    search_candidates = sheet_records(workbook, "Search_Candidates")
    selected_configs = sheet_records(workbook, "Selected_Config")
    extra_json = sheet_records(workbook, "Extra_JSON_Rows")
    states = sheet_records(workbook, "State_JSON")
    csv_rows = sheet_records(workbook, "All_CSV_Rows")

    raw_fields = [
        "category",
        "dataset",
        "method",
        "ablation",
        "scenario",
        "few_shot_ratio",
        "seed",
        "test_auc",
        "test_ap",
        "test_mrr",
        "test_hits@10",
        "test_tgb_mrr",
        "test_tgb_hits@10",
        "test_ndcg",
        "val_auc",
        "val_ap",
        "val_mrr",
        "val_hits@10",
        "val_tgb_mrr",
        "val_tgb_hits@10",
        "val_ndcg",
    ]
    raw_keys = {result_key(row, raw_fields) for row in raw}
    partial_unique = [row for row in partial if result_key(row, raw_fields) not in raw_keys]

    search_fields = [
        "candidate",
        "seed",
        "train_loss",
        "val_auc",
        "val_ap",
        "val_mrr",
        "val_hits@10",
        "val_tgb_mrr",
        "val_tgb_hits@10",
        "val_ndcg",
        "error",
        "error_type",
        "status",
    ]
    search_keys = {result_key(row, search_fields) for row in search_runs}
    search_json_unique = [
        row for row in search_json if result_key(row, search_fields) not in search_keys
    ]

    # Extra_JSON_Rows uses file-oriented category names (for example,
    # formal_final), while All_Results_Raw uses experiment-oriented names
    # (formal_main).  Category is therefore not part of cross-sheet identity.
    scientific_fields = raw_fields[1:]
    scientific_raw_keys = {result_key(row, scientific_fields) for row in raw}
    extra_metric_rows = [
        row
        for row in extra_json
        if row.get("json_section") == "rows"
        and any(row.get(key) is not None for key in ("test_auc", "test_ap", "val_auc", "val_ap"))
    ]
    extra_unique = [
        row
        for row in extra_metric_rows
        if result_key(row, scientific_fields) not in scientific_raw_keys
    ]
    seen_extra: set[tuple[Any, ...]] = set()
    extra_unique = [
        row
        for row in extra_unique
        if not (
            result_key(row, scientific_fields) in seen_extra
            or seen_extra.add(result_key(row, scientific_fields))
        )
    ]

    counts = {
        "summaries": len(summaries),
        "raw": len(raw),
        "partial": len(partial),
        "partial_unique": len(partial_unique),
        "search_runs": len(search_runs),
        "search_summaries": len(search_summaries),
        "search_json": len(search_json),
        "search_json_unique": len(search_json_unique),
        "search_candidates": len(search_candidates),
        "selected_configs": len(selected_configs),
        "extra_metric_rows": len(extra_metric_rows),
        "extra_json": len(extra_json),
        "extra_unique": len(extra_unique),
        "states": len(states),
        "csv_rows": len(csv_rows),
    }

    lines: list[str] = [
        BEGIN,
        "",
        "## 完整实验结果附录",
        "",
        "本附录以 `EXPERIMENT_RESULTS.xlsx` 为结构化来源，补齐正文未逐行展示的结果。"
        "同一实验可能同时写入合并文件、seed 分片、断点文件和单实验 JSON；这里保留"
        " `All_Results_Raw` 的全部来源行，并对其他格式只补其独有结果，避免把文件副本误称为独立实验。",
        "",
        "指标列严格区分普通采样排名的 `MRR/H@K` 与 TGB 官方协议的 `TGB-MRR/TGB-H@10`；"
        "空值以 `-` 表示，不能与 0 混同。`invalid_history` 和 `smoke` 仅为审计记录，不参与论文结论。",
        "",
        "### 覆盖核对",
        "",
    ]

    coverage_rows = [
        {"sheet": "All_Summaries", "source": len(summaries), "included": len(summaries), "policy": "全部汇总行"},
        {"sheet": "All_Results_Raw", "source": len(raw), "included": len(raw), "policy": "全部逐 seed/source 行"},
        {"sheet": "Partial_Raw", "source": len(partial), "included": len(partial_unique), "policy": "仅补 All_Results_Raw 中不存在的行"},
        {"sheet": "Search_Summary", "source": len(search_summaries), "included": len(search_summaries), "policy": "已在正文完整展示"},
        {"sheet": "Search_Runs", "source": len(search_runs), "included": len(search_runs), "policy": "全部 trial/seed 行，含失败状态"},
        {"sheet": "Search_JSON_Rows", "source": len(search_json), "included": len(search_json_unique), "policy": "仅补 Search_Runs 中不存在的行"},
        {"sheet": "Search_Candidates", "source": len(search_candidates), "included": len(search_candidates), "policy": "全部候选汇总"},
        {"sheet": "Selected_Config", "source": len(selected_configs), "included": len(selected_configs), "policy": "全部配置，展示关键字段"},
        {"sheet": "Extra_JSON_Rows", "source": len(extra_json), "included": len(extra_unique), "policy": f"{len(extra_metric_rows)} 行带指标；仅补独有指标行"},
        {"sheet": "State_JSON", "source": len(states), "included": len(states), "policy": "全部运行状态行"},
        {"sheet": "All_CSV_Rows", "source": len(csv_rows), "included": 0, "policy": "CSV 序列化投影；结果已由汇总、对比和候选表覆盖"},
    ]
    lines.extend(render_table(coverage_rows, [("工作表", "sheet"), ("源行数", "source"), ("正文独立展示行数", "included"), ("去重口径", "policy")]))

    summary_columns: list[Column] = [
        ("category", "category"),
        ("experiment", "experiment_dir"),
        ("dataset/group", nonempty("dataset", "group")),
        ("method/signature", nonempty("method", "model_signature")),
        ("ablation", "ablation"),
        ("scenario", "scenario"),
        ("ratio", "few_shot_ratio"),
        ("test AUC", "test_auc_mean"),
        ("test AP", "test_ap_mean"),
        ("test MRR", "test_mrr_mean"),
        ("test H@10", "test_hits@10_mean"),
        ("TGB-MRR", "test_tgb_mrr_mean"),
        ("TGB-H@10", "test_tgb_hits@10_mean"),
        ("test NDCG", "test_ndcg_mean"),
        ("mean val score", "mean_score"),
        ("mean val AUC", "mean_val_auc"),
        ("mean val AP", "mean_val_ap"),
        ("mean val MRR", "mean_val_mrr"),
        ("mean val H@10", "mean_val_hits@10"),
        ("mean val NDCG", "mean_val_ndcg"),
    ]
    lines.extend(["", "### 全部汇总级结果（255 行）", ""])
    lines.extend(details("展开 All_Summaries 全部结果", render_table(summaries, summary_columns)))

    raw_columns: list[Column] = [
        ("source experiment", "experiment_dir"),
        ("dataset", "dataset"),
        ("method", "method"),
        ("ablation", "ablation"),
        ("scenario", "scenario"),
        ("ratio", "few_shot_ratio"),
        ("seed", "seed"),
        ("test AUC", "test_auc"),
        ("test AP", "test_ap"),
        ("test MRR", "test_mrr"),
        ("test H@1", "test_hits@1"),
        ("test H@10", "test_hits@10"),
        ("test H@50", "test_hits@50"),
        ("test H@100", "test_hits@100"),
        ("test NDCG", "test_ndcg"),
        ("TGB-MRR", "test_tgb_mrr"),
        ("TGB-H@10", "test_tgb_hits@10"),
        ("val AUC", "val_auc"),
        ("val AP", "val_ap"),
        ("val MRR", "val_mrr"),
        ("val H@10", "val_hits@10"),
        ("val NDCG", "val_ndcg"),
        ("val TGB-MRR", "val_tgb_mrr"),
        ("val TGB-H@10", "val_tgb_hits@10"),
        ("params", "params_total"),
        ("FLOPs", "rough_forward_flops"),
        ("max GPU MB", "max_allocated_mb"),
        ("wall sec", "wall_time_sec"),
    ]
    lines.extend(["", "### 全部逐 seed 结果（1303 行）", ""])
    lines.append(
        "这些行保留合并结果与 seed 分片的来源级记录，因此同一科学结果可能出现两次；"
        "保留重复项是为了与工作簿逐行可追溯，而不是增加样本量。"
    )
    raw_by_category: dict[str, list[dict[str, Any]]] = {}
    for row in raw:
        raw_by_category.setdefault(str(row.get("category") or "uncategorized"), []).append(row)
    for category in sorted(raw_by_category):
        rows = raw_by_category[category]
        lines.extend(["", *details(f"{category}: {len(rows)} 行", render_table(rows, raw_columns))])

    search_run_columns: list[Column] = [
        ("category", "category"),
        ("search", "experiment_dir"),
        ("candidate", "candidate"),
        ("seed", "seed"),
        ("status", "status"),
        ("composite", "composite"),
        ("val AUC", "val_auc"),
        ("val AP", "val_ap"),
        ("val MRR", "val_mrr"),
        ("val H@1", "val_hits@1"),
        ("val H@10", "val_hits@10"),
        ("val H@50", "val_hits@50"),
        ("val H@100", "val_hits@100"),
        ("val NDCG", "val_ndcg"),
        ("val TGB-MRR", "val_tgb_mrr"),
        ("val TGB-H@10", "val_tgb_hits@10"),
        ("error type", "error_type"),
        ("error", "error"),
    ]
    lines.extend(["", "### 全部调参 trial 结果（702 行）", ""])
    lines.extend(details("展开 Search_Runs 全部结果", render_table(search_runs, search_run_columns)))

    candidate_columns: list[Column] = [
        ("category", "category"),
        ("search", "experiment_dir"),
        ("candidate", "candidate"),
        ("seeds", "seeds_completed"),
        ("passes minima", "passes_metric_minima"),
        ("score", "score"),
        ("val AUC", "means_val_auc"),
        ("val AP", "means_val_ap"),
        ("val MRR", "means_val_mrr"),
        ("val H@10", "means_val_hits@10"),
        ("val H@50", "means_val_hits@50"),
        ("val H@100", "means_val_hits@100"),
        ("val NDCG", "means_val_ndcg"),
        ("val TGB-MRR", "means_val_tgb_mrr"),
        ("val TGB-H@10", "means_val_tgb_hits@10"),
    ]
    lines.extend(["", "### 全部调参候选汇总（401 行）", ""])
    lines.extend(details("展开 Search_Candidates 全部结果", render_table(search_candidates, candidate_columns)))

    if search_json_unique:
        lines.extend(["", f"### Search JSON 独有结果（{len(search_json_unique)} 行）", ""])
        lines.append(
            f"`Search_JSON_Rows` 共 {len(search_json)} 行，其中 {len(search_json) - len(search_json_unique)} 行已在 "
            f"`Search_Runs` 出现；下表仅列出其余 {len(search_json_unique)} 行。"
        )
        lines.extend(["", *details("展开 Search_JSON_Rows 独有结果", render_table(search_json_unique, search_run_columns))])

    config_columns: list[Column] = [
        ("category", "category"),
        ("experiment", "experiment_dir"),
        ("dataset", "data_name"),
        ("scenario", "scenario_name"),
        ("device", "device"),
        ("hidden", "model_hidden_dim"),
        ("adapter rank", "model_adapter_rank"),
        ("relation rank", "model_relation_rank"),
        ("temporal encoder", "model_temporal_encoder"),
        ("time encoder", "model_time_encoder"),
        ("cross", "model_use_cross"),
        ("gate", "model_use_gate"),
        ("history prior", "model_use_history_prior"),
        ("semantic", "model_use_sem"),
        ("structure", "model_use_struct"),
        ("variational", "model_use_variational"),
        ("lr", "train_lr"),
        ("epochs", "train_epochs"),
        ("patience", "train_patience"),
        ("train negatives", "train_num_neg_train"),
        ("negative mode", "train_negative_mode_train"),
        ("rank loss", "train_rank_loss_type"),
    ]
    lines.extend(["", "### 全部选中配置（81 行）", ""])
    lines.append("下表展示决定模型和训练行为的关键字段；108 个完整配置字段仍保留在同名 Excel 工作表。")
    lines.extend(["", *details("展开 Selected_Config 全部配置", render_table(selected_configs, config_columns))])

    extra_columns: list[Column] = [
        ("source", "source_file"),
        ("name", "name"),
        ("val AUC", "val_auc"),
        ("val AP", "val_ap"),
        ("val MRR", "val_mrr"),
        ("val H@1", "val_hits@1"),
        ("val H@10", "val_hits@10"),
        ("val H@50", "val_hits@50"),
        ("val H@100", "val_hits@100"),
        ("val NDCG", "val_ndcg"),
    ]
    lines.extend(["", f"### 其他 JSON 独有指标结果（{len(extra_unique)} 行）", ""])
    lines.extend(render_table(extra_unique, extra_columns))

    state_columns: list[Column] = [
        ("source", "source_file"),
        ("dataset", "dataset"),
        ("section", "state_section"),
        ("index", "state_index"),
        ("score", "score"),
        ("wins", "wins"),
        ("candidate", nonempty(
            "candidate_listwise_yago_diag_probe.best_candidate",
            "global_relpop_rep_enhanced_probe.best_candidate",
            "history_popularity_quickscan.best_candidate",
            "history_popularity_tinysearch.best_candidate",
            "rankloss_micro_fastscan.best_candidate",
            "rankloss_uniform_probe.best_candidate",
            "struct_history_fastscan.best_candidate",
            "struct_residual_fastscan.best_candidate",
        )),
        ("note", "note"),
        ("path", "path"),
    ]
    lines.extend(["", "### 全部运行状态记录（134 行）", ""])
    lines.append("状态行用于追踪优化任务是否完成，不作为独立测试集结果或独立 seed。")
    lines.extend(["", *details("展开 State_JSON 全部状态", render_table(states, state_columns))])

    if partial_unique:
        lines.extend(["", f"### Partial_Raw 独有结果（{len(partial_unique)} 行）", ""])
        lines.extend(render_table(partial_unique, raw_columns))
    else:
        lines.extend([
            "",
            "### 断点结果去重说明",
            "",
            f"`Partial_Raw` 的 {len(partial)} 行按数据集、方法、消融、场景、seed、测试和验证指标核对后，"
            "均已存在于 `All_Results_Raw`，因此不重复抄录。",
        ])

    lines.extend(["", END])
    return "\n".join(lines), counts


def update_report(appendix: str, counts: dict[str, int]) -> None:
    text = REPORT.read_text(encoding="utf-8")
    text = text.replace("# EXPERIMENT_RESULTS", "# SSP-TGFM 实验结果总汇", 1)
    text = text.replace("生成日期：2026-07-22", f"生成日期：{date.today().isoformat()}", 1)
    text = text.replace("## 正式主结果\n###", "## 正式主结果\n\n###")
    text = text.replace("\n## Excel 表单", "\n\n## Excel 表单")

    new_sentence = (
        f"- Markdown 已直接覆盖：汇总结果 {counts['summaries']} 行、逐 seed/source 结果 {counts['raw']} 行、"
        f"调参 trial {counts['search_runs']} 行、候选汇总 {counts['search_candidates']} 行、"
        f"选中配置 {counts['selected_configs']} 行、Search JSON 独有结果 {counts['search_json_unique']} 行、"
        f"其他 JSON 独有指标 {counts['extra_unique']} 行和运行状态 {counts['states']} 行。"
    )
    report_lines = text.splitlines()
    for index, line in enumerate(report_lines):
        if line.startswith("- 结构化结果已展开到 Excel：") or line.startswith("- Markdown 已直接覆盖："):
            report_lines[index] = new_sentence
        elif line.startswith("- `invalid_struct_field_order` 结果只保留在 Excel"):
            report_lines[index] = (
                "- `invalid_struct_field_order` 结果在完整附录和 Excel 的 `Invalid_History` 中留档，"
                "明确标记作废且不纳入主结论。"
            )
    text = "\n".join(report_lines) + "\n"

    if BEGIN in text and END in text:
        before, rest = text.split(BEGIN, 1)
        _, after = rest.split(END, 1)
        text = before.rstrip() + "\n\n" + appendix + "\n\n" + after.lstrip("\n")
    else:
        marker = "## Excel 表单"
        if marker not in text:
            raise RuntimeError(f"Insertion marker not found in {REPORT}")
        before, after = text.split(marker, 1)
        text = before.rstrip() + "\n\n" + appendix + "\n\n" + marker + after

    # Normalize heading spacing so tables and HTML details render consistently.
    source_lines = text.rstrip().splitlines()
    normalized: list[str] = []
    for index, line in enumerate(source_lines):
        is_heading = line.startswith("#")
        if is_heading and normalized and normalized[-1] != "":
            normalized.append("")
        normalized.append(line)
        if is_heading and index + 1 < len(source_lines) and source_lines[index + 1] != "":
            normalized.append("")
    REPORT.write_text("\n".join(normalized).rstrip() + "\n", encoding="utf-8")


def main() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    appendix, counts = build_appendix(workbook)
    update_report(appendix, counts)
    print(json.dumps(counts, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
