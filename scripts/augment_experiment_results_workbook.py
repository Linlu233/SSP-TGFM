#!/usr/bin/env python3
"""Add full result-artifact coverage sheets to EXPERIMENT_RESULTS.xlsx.

The first workbook generation focused on curated metric tables. This pass keeps
those sheets and adds traceability for every structured and run-artifact file
under results/.
"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "results"
WORKBOOK = ROOT / "EXPERIMENT_RESULTS.xlsx"

STANDARD_JSON_NAMES = {
    "all_results.json",
    "search_results.json",
    "selected_config.json",
}
STATE_JSON_NAMES = {"state.json", "state_probe.json"}
TEXT_SUFFIXES = {".log", ".txt", ".md", ".out", ".pid"}
ILLEGAL_EXCEL_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
MAX_CELL_CHARS = 30000


def safe_cell(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    else:
        value = str(value)
    value = ILLEGAL_EXCEL_RE.sub("", value)
    if len(value) > MAX_CELL_CHARS:
        return value[:MAX_CELL_CHARS] + "...[truncated]"
    return value


def flatten(value: Any, prefix: str = "", max_depth: int = 5) -> dict[str, Any]:
    out: dict[str, Any] = {}

    def walk(obj: Any, key: str, depth: int) -> None:
        if depth > max_depth:
            out[key] = obj
            return
        if isinstance(obj, dict):
            if not obj:
                out[key] = {}
            for raw_k, raw_v in obj.items():
                clean_k = str(raw_k).replace(".", "_")
                next_key = f"{key}.{clean_k}" if key else clean_k
                walk(raw_v, next_key, depth + 1)
        elif isinstance(obj, list):
            out[key] = obj
        else:
            out[key] = obj

    walk(value, prefix, 0)
    return out


def json_load(path: Path) -> tuple[Any | None, str | None]:
    try:
        return json.loads(path.read_text(encoding="utf-8")), None
    except Exception as exc:  # pragma: no cover - reporting script
        return None, repr(exc)


def count_csv_rows(path: Path) -> int | None:
    try:
        with path.open(newline="", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            next(reader, None)
            return sum(1 for _ in reader)
    except Exception:
        return None


def count_json_rows(path: Path) -> int | None:
    data, err = json_load(path)
    if err is not None:
        return None
    if isinstance(data, list):
        return len(data)
    if isinstance(data, dict):
        if isinstance(data.get("rows"), list):
            return len(data["rows"])
        if isinstance(data.get("candidate_summaries"), list):
            return len(data["candidate_summaries"])
        if isinstance(data.get("prepared"), list):
            return len(data["prepared"])
        if isinstance(data.get("datasets"), dict):
            return len(data["datasets"])
    return 1


def count_jsonl_rows(path: Path) -> int | None:
    try:
        with path.open(encoding="utf-8") as fh:
            return sum(1 for line in fh if line.strip())
    except Exception:
        return None


def row_count(path: Path) -> int | None:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return count_csv_rows(path)
    if suffix == ".json":
        return count_json_rows(path)
    if suffix == ".jsonl":
        return count_jsonl_rows(path)
    return None


def result_category(path: Path) -> str:
    rel = path.relative_to(ROOT).as_posix()
    name = path.name
    if "invalid_" in rel:
        return "invalid_history"
    if "final_unified_s15sf02_official" in rel:
        return "formal_final"
    if "global_model_formal_auto" in rel:
        return "formal_global"
    if "ablation_unified_s15sf02_official" in rel:
        return "ablation"
    if "baseline" in rel:
        return "baseline"
    if "search" in rel:
        return "search"
    if "synthetic" in rel:
        return "synthetic"
    if "optimization" in rel or "auto_target" in rel or "global_" in rel:
        return "optimization_or_candidate"
    if "tgb_icews" in rel:
        return "excluded_tgb_icews"
    if name in STATE_JSON_NAMES or name.endswith(".status.json"):
        return "run_state"
    return "other"


def json_kind(path: Path, data: Any) -> str:
    name = path.name
    if name in STATE_JSON_NAMES:
        return "state"
    if name == "manifest.json":
        return "manifest"
    if name.endswith(".status.json"):
        return "status"
    if name.startswith("STOP_NOTE"):
        return "stop_note"
    if name.endswith("_selection.json") or "validation_" in name:
        return "selection_or_validation"
    if name == "explanations.json":
        return "explanations"
    if "history_prior" in name:
        return "history_prior"
    if isinstance(data, dict) and isinstance(data.get("rows"), list):
        return "single_run_result"
    if isinstance(data, list):
        return "json_list"
    return "json_metadata"


def rel_path(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def mtime_utc(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()


def inventory_rows(files: list[Path]) -> list[dict[str, Any]]:
    rows = []
    for path in files:
        suffix = path.suffix.lower()
        parsed = suffix in {".json", ".jsonl", ".csv"}
        rows.append(
            {
                "source_file": rel_path(path),
                "directory": path.parent.relative_to(ROOT).as_posix(),
                "filename": path.name,
                "suffix": suffix or "<none>",
                "category": result_category(path),
                "size_bytes": path.stat().st_size,
                "mtime_utc": mtime_utc(path),
                "row_count_if_structured": row_count(path),
                "structured": parsed,
                "invalid_history": "invalid_" in rel_path(path),
                "tgb_icews_related": "tgb_icews" in rel_path(path),
            }
        )
    return rows


def build_index(inv: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in inv:
        key = (row["category"], row["filename"] if row["structured"] else row["suffix"])
        item = grouped.setdefault(
            key,
            {
                "category": row["category"],
                "file_type": key[1],
                "file_count": 0,
                "row_count": 0,
                "total_size_bytes": 0,
            },
        )
        item["file_count"] += 1
        item["row_count"] += row["row_count_if_structured"] or 0
        item["total_size_bytes"] += row["size_bytes"]
    return [grouped[k] for k in sorted(grouped)]


def build_extra_json(files: list[Path]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    summary: list[dict[str, Any]] = []
    expanded: list[dict[str, Any]] = []

    for path in files:
        data, err = json_load(path)
        source = rel_path(path)
        base = {
            "source_file": source,
            "directory": path.parent.relative_to(ROOT).as_posix(),
            "filename": path.name,
            "category": result_category(path),
            "size_bytes": path.stat().st_size,
            "mtime_utc": mtime_utc(path),
        }
        if err is not None:
            summary.append({**base, "json_kind": "parse_error", "parse_error": err})
            continue

        if isinstance(data, dict):
            keys = sorted(data.keys())
            rows_count = len(data.get("rows", [])) if isinstance(data.get("rows"), list) else None
            summary_keys = sorted(data.get("summary", {}).keys()) if isinstance(data.get("summary"), dict) else []
            config_keys = sorted(data.get("config", {}).keys()) if isinstance(data.get("config"), dict) else []
            summary.append(
                {
                    **base,
                    "json_kind": json_kind(path, data),
                    "json_type": "dict",
                    "top_level_keys": ", ".join(keys),
                    "rows_count": rows_count,
                    "summary_keys": ", ".join(summary_keys),
                    "config_keys": ", ".join(config_keys),
                }
            )
            add_expanded_json_rows(expanded, base, data)
        elif isinstance(data, list):
            summary.append(
                {
                    **base,
                    "json_kind": json_kind(path, data),
                    "json_type": "list",
                    "top_level_keys": "",
                    "rows_count": len(data),
                }
            )
            for idx, item in enumerate(data):
                row = {**base, "json_section": "list", "json_index": idx}
                row.update(flatten(item) if isinstance(item, dict) else {"value": item})
                expanded.append(row)
        else:
            summary.append(
                {
                    **base,
                    "json_kind": json_kind(path, data),
                    "json_type": type(data).__name__,
                    "top_level_keys": "",
                    "rows_count": 1,
                }
            )
            expanded.append({**base, "json_section": "scalar", "value": data})
    return summary, expanded


def add_expanded_json_rows(out: list[dict[str, Any]], base: dict[str, Any], data: dict[str, Any]) -> None:
    config = flatten(data.get("config", {}), "config") if isinstance(data.get("config"), dict) else {}
    scenario = data.get("scenario")
    common = {**base, **config}
    if scenario is not None:
        common["scenario"] = scenario

    rows = data.get("rows")
    if isinstance(rows, list):
        for idx, item in enumerate(rows):
            row = {**common, "json_section": "rows", "json_index": idx}
            row.update(flatten(item) if isinstance(item, dict) else {"value": item})
            out.append(row)

    summary = data.get("summary")
    if isinstance(summary, dict):
        row = {**common, "json_section": "summary", "json_index": 0}
        row.update(flatten(summary, "summary"))
        out.append(row)

    for section in ("prepared", "blocked_ablations", "candidate_summaries"):
        value = data.get(section)
        if isinstance(value, list):
            for idx, item in enumerate(value):
                row = {**common, "json_section": section, "json_index": idx}
                row.update(flatten(item) if isinstance(item, dict) else {"value": item})
                out.append(row)

    datasets = data.get("datasets")
    if isinstance(datasets, dict):
        shared = flatten(data.get("shared_model_config", {}), "shared_model_config")
        for dataset, value in datasets.items():
            row = {**common, **shared, "json_section": "datasets", "dataset_key": dataset}
            row.update(flatten(value, "dataset_state") if isinstance(value, dict) else {"value": value})
            out.append(row)

    handled = any(
        isinstance(data.get(section), expected)
        for section, expected in (
            ("rows", list),
            ("summary", dict),
            ("prepared", list),
            ("blocked_ablations", list),
            ("candidate_summaries", list),
            ("datasets", dict),
        )
    )
    if not handled:
        row = {**common, "json_section": "metadata", "json_index": 0}
        row.update(flatten(data))
        out.append(row)


def build_state_json(files: list[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in files:
        data, err = json_load(path)
        base = {
            "source_file": rel_path(path),
            "directory": path.parent.relative_to(ROOT).as_posix(),
            "filename": path.name,
            "category": result_category(path),
            "size_bytes": path.stat().st_size,
            "mtime_utc": mtime_utc(path),
        }
        if err is not None:
            rows.append({**base, "state_section": "parse_error", "parse_error": err})
            continue
        if isinstance(data, dict) and isinstance(data.get("datasets"), dict):
            for dataset, dataset_state in data["datasets"].items():
                if isinstance(dataset_state, dict):
                    for section, value in dataset_state.items():
                        if isinstance(value, list):
                            for idx, item in enumerate(value):
                                row = {**base, "dataset": dataset, "state_section": section, "state_index": idx}
                                row.update(flatten(item) if isinstance(item, dict) else {"value": item})
                                rows.append(row)
                        else:
                            row = {**base, "dataset": dataset, "state_section": section, "state_index": 0}
                            row.update(flatten(value) if isinstance(value, dict) else {"value": value})
                            rows.append(row)
                else:
                    rows.append({**base, "dataset": dataset, "state_section": "datasets", "value": dataset_state})
        else:
            row = {**base, "state_section": "metadata", "state_index": 0}
            row.update(flatten(data) if isinstance(data, dict) else {"value": data})
            rows.append(row)
    return rows


def build_csv_rows(files: list[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in files:
        source = rel_path(path)
        try:
            with path.open(newline="", encoding="utf-8") as fh:
                reader = csv.DictReader(fh)
                for idx, item in enumerate(reader):
                    rows.append(
                        {
                            "source_file": source,
                            "directory": path.parent.relative_to(ROOT).as_posix(),
                            "filename": path.name,
                            "category": result_category(path),
                            "csv_row_index": idx,
                            **item,
                        }
                    )
        except Exception as exc:  # pragma: no cover - reporting script
            rows.append(
                {
                    "source_file": source,
                    "directory": path.parent.relative_to(ROOT).as_posix(),
                    "filename": path.name,
                    "category": result_category(path),
                    "parse_error": repr(exc),
                }
            )
    return rows


def read_text_edge(path: Path, from_end: bool = False, chunk_size: int = 4096) -> str:
    if path.stat().st_size == 0:
        return ""
    mode = "rb"
    with path.open(mode) as fh:
        if from_end:
            fh.seek(max(0, path.stat().st_size - chunk_size))
        data = fh.read(chunk_size)
    text = data.decode("utf-8", errors="replace")
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return ""
    return lines[-1] if from_end else lines[0]


def build_run_artifacts(files: list[Path]) -> list[dict[str, Any]]:
    rows = []
    for path in files:
        rows.append(
            {
                "source_file": rel_path(path),
                "directory": path.parent.relative_to(ROOT).as_posix(),
                "filename": path.name,
                "suffix": path.suffix.lower() or "<none>",
                "category": result_category(path),
                "size_bytes": path.stat().st_size,
                "mtime_utc": mtime_utc(path),
                "first_nonempty_snippet": read_text_edge(path, from_end=False),
                "last_nonempty_snippet": read_text_edge(path, from_end=True),
            }
        )
    return rows


def write_sheet(wb, name: str, rows: list[dict[str, Any]], preferred: list[str] | None = None) -> None:
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    preferred = preferred or []
    if not rows:
        ws.append(["note"])
        ws.append(["No rows"])
        return

    keys = set()
    for row in rows:
        keys.update(row.keys())
    columns = [key for key in preferred if key in keys]
    columns.extend(sorted(keys.difference(columns)))
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append([safe_cell(row.get(column)) for column in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, column in enumerate(columns, start=1):
        width = min(max(len(str(column)) + 2, 12), 42)
        ws.column_dimensions[get_column_letter(idx)].width = width


def rewrite_sheet_guide(wb) -> None:
    guide = {
        "Sheet_Guide": "本表：说明每个 sheet 的内容。",
        "Index": "全量 results/ 文件按类别和文件类型聚合统计。",
        "RootDocs": "项目根目录文档审阅。",
        "Formal_Final": "final_unified_s15sf02_official，建议作为主正式结果线。",
        "Formal_Global": "global_model_formal_auto，统一结构候选/备选正式线。",
        "Baselines": "已跑快速强基线与可训练强基线。",
        "Formal_vs_Base": "Formal_Final 与已跑强基线逐指标对比。",
        "Ablations": "统一正式模型的消融汇总。",
        "Ablation_Best": "每个数据集每个主指标上 full 与最佳 ablation 对比。",
        "Synthetic_Core": "未标记作废的合成/迁移协议结果。",
        "Synthetic_Strong": "stronggate_v4_formal 候选/优化结果。",
        "Invalid_History": "已标记 invalid_struct_field_order 的历史作废结果。",
        "Search_Summary": "调参搜索 search_results.json 摘要。",
        "Search_Runs": "调参搜索 partial_search_results.jsonl 全部候选行。",
        "Search_JSON_Rows": "search_results.json 内 rows 展平。",
        "Search_Candidates": "search_results.json 中 candidate_summaries。",
        "Selected_Config": "所有 selected_config.json 展平后的配置。",
        "Candidate_Formal": "其他候选正式/strict_v2/best 汇总。",
        "All_Summaries": "所有 summary.csv 与 global summary 合并。",
        "All_Results_Raw": "所有 all_results.json 的 seed/raw rows。",
        "Partial_Raw": "所有 partial_results.jsonl 的断点 seed/raw rows。",
        "Inventory": "results/ 下全量文件索引，含结构化文件、日志、pid/out 等运行产物。",
        "Extra_JSON_Summary": "非标准命名 JSON 的逐文件摘要，包括单实验 JSON、manifest、selection、status、STOP_NOTE、validation 等。",
        "Extra_JSON_Rows": "非标准命名 JSON 中可展开内容的扁平行。",
        "State_JSON": "state.json/state_probe.json 自动优化状态元数据展开。",
        "All_CSV_Rows": "results/ 下所有 CSV 的原始行合并。",
        "Run_Artifacts": "日志、txt、md、out、pid 等非结构化运行产物索引与首尾片段。",
    }
    rows = [{"sheet": sheet, "content": guide.get(sheet, "")} for sheet in wb.sheetnames]
    write_sheet(wb, "Sheet_Guide", rows, ["sheet", "content"])
    wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index("Sheet_Guide")))


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"Missing workbook: {WORKBOOK}")
    if not RESULTS.exists():
        raise SystemExit(f"Missing results directory: {RESULTS}")

    all_files = sorted(path for path in RESULTS.rglob("*") if path.is_file())
    json_files = [path for path in all_files if path.suffix.lower() == ".json"]
    csv_files = [path for path in all_files if path.suffix.lower() == ".csv"]
    state_json_files = [path for path in json_files if path.name in STATE_JSON_NAMES]
    extra_json_files = [
        path
        for path in json_files
        if path.name not in STANDARD_JSON_NAMES and path.name not in STATE_JSON_NAMES
    ]
    run_artifact_files = [path for path in all_files if path.suffix.lower() in TEXT_SUFFIXES]

    inv = inventory_rows(all_files)
    index = build_index(inv)
    extra_summary, extra_rows = build_extra_json(extra_json_files)
    state_rows = build_state_json(state_json_files)
    csv_rows = build_csv_rows(csv_files)
    run_rows = build_run_artifacts(run_artifact_files)

    wb = load_workbook(WORKBOOK)
    write_sheet(
        wb,
        "Inventory",
        inv,
        [
            "source_file",
            "directory",
            "filename",
            "suffix",
            "category",
            "size_bytes",
            "mtime_utc",
            "row_count_if_structured",
            "structured",
            "invalid_history",
            "tgb_icews_related",
        ],
    )
    write_sheet(wb, "Index", index, ["category", "file_type", "file_count", "row_count", "total_size_bytes"])
    write_sheet(
        wb,
        "Extra_JSON_Summary",
        extra_summary,
        [
            "source_file",
            "directory",
            "filename",
            "category",
            "json_kind",
            "json_type",
            "rows_count",
            "top_level_keys",
            "summary_keys",
            "config_keys",
            "size_bytes",
            "mtime_utc",
        ],
    )
    write_sheet(
        wb,
        "Extra_JSON_Rows",
        extra_rows,
        ["source_file", "directory", "filename", "category", "json_section", "json_index", "dataset", "dataset_key", "scenario"],
    )
    write_sheet(
        wb,
        "State_JSON",
        state_rows,
        ["source_file", "directory", "filename", "category", "dataset", "state_section", "state_index"],
    )
    write_sheet(
        wb,
        "All_CSV_Rows",
        csv_rows,
        ["source_file", "directory", "filename", "category", "csv_row_index", "dataset", "method", "ablation", "scenario"],
    )
    write_sheet(
        wb,
        "Run_Artifacts",
        run_rows,
        [
            "source_file",
            "directory",
            "filename",
            "suffix",
            "category",
            "size_bytes",
            "mtime_utc",
            "first_nonempty_snippet",
            "last_nonempty_snippet",
        ],
    )
    rewrite_sheet_guide(wb)
    wb.save(WORKBOOK)

    suffix_counts = Counter(path.suffix.lower() or "<none>" for path in all_files)
    print(f"workbook={WORKBOOK}")
    print(f"all_files={len(all_files)}")
    print("suffix_counts=" + json.dumps(dict(sorted(suffix_counts.items())), ensure_ascii=False))
    print(f"extra_json_summary={len(extra_summary)}")
    print(f"extra_json_rows={len(extra_rows)}")
    print(f"state_json_rows={len(state_rows)}")
    print(f"all_csv_rows={len(csv_rows)}")
    print(f"run_artifacts={len(run_rows)}")


if __name__ == "__main__":
    main()
